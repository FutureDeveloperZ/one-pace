#!/usr/bin/env python3
"""Downloads the One Pace Episode Guide spreadsheet and extracts episode data to episodes.json."""

import json
import re
import sys
import urllib.request
import tempfile
import os

SPREADSHEET_ID = "1HQRMJgu_zArp-sLnvFMDzOyjdsht87eFLECxMK858lA"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "..", "data", "episodes.json")


def download_spreadsheet():
    """Download the spreadsheet from Google Sheets."""
    print(f"Downloading spreadsheet from Google Sheets...")
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    urllib.request.urlretrieve(EXPORT_URL, tmp.name)
    print(f"Downloaded to {tmp.name}")
    return tmp.name


def parse_episodes(xlsx_path):
    """Parse episodes from the spreadsheet."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    all_episodes = []
    arc_overview = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Arc Overview":
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                arc_name = row[1]
                if not arc_name:
                    continue
                num_pace = row[7]
                try:
                    num_pace = int(num_pace) if num_pace else 0
                except (ValueError, TypeError):
                    num_pace = 0
                arc_overview[str(arc_name).strip()] = {
                    "paceEpisodeCount": num_pace
                }
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        for row in rows[1:]:
            pace_ep = row[1]
            chapters = row[2]
            episodes = row[3]
            length = row[5]
            if pace_ep is None or str(pace_ep).strip() == "":
                continue

            length_min = 0
            if length:
                if hasattr(length, "hour"):
                    length_min = length.hour * 60 + length.minute + length.second / 60
                elif isinstance(length, (int, float)):
                    length_min = float(length)

            chapters_str = str(chapters).strip() if chapters else ""
            chapter_list = []
            if chapters_str:
                for part in chapters_str.replace(" ", "").split(","):
                    part = part.strip()
                    if "-" in part:
                        a, b = part.split("-", 1)
                        chapter_list.append({"from": a.strip(), "to": b.strip()})
                    elif part:
                        chapter_list.append({"from": part, "to": part})

            episodes_str = str(episodes).strip() if episodes else ""
            op_start = None
            op_end = None
            if episodes_str:
                cleaned = episodes_str.replace("Ep.", "").replace("Episode of", "").strip()
                cleaned = cleaned.replace("\u2014", "-").replace("\u2013", "-")
                cleaned = re.sub(r"[^0-9,\-]", " ", cleaned)
                numbers = [int(x) for x in re.findall(r"\d+", cleaned)]
                if numbers:
                    op_start = min(numbers)
                    op_end = max(numbers)

            all_episodes.append({
                "arc": sheet_name.strip(),
                "paceEpisode": str(pace_ep).strip(),
                "chapters": chapter_list,
                "chaptersRaw": chapters_str,
                "opStart": op_start,
                "opEnd": op_end,
                "opEpisodesRaw": episodes_str,
                "lengthMinutes": round(length_min, 1)
            })

    wb.close()
    return arc_overview, all_episodes


def main():
    xlsx_path = None
    try:
        xlsx_path = download_spreadsheet()
        arcs, episodes = parse_episodes(xlsx_path)

        data = {
            "arcs": arcs,
            "episodes": episodes
        }

        os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
        with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"Extracted {len(episodes)} episodes across {len(arcs)} arcs")
        print(f"Written to {OUTPUT_PATH}")

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        if xlsx_path and os.path.exists(xlsx_path):
            os.unlink(xlsx_path)


if __name__ == "__main__":
    main()
